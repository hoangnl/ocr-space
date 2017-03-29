from ocr import ocr_space_file
import json
import openpyxl
import os
import shutil

file_name = 'example_image.png'
test_file = ocr_space_file(filename=file_name, language='pol')
json_obj = json.loads(test_file)
actual_result = ""
for key in json_obj["ParsedResults"][0]["TextOverlay"]["Lines"]:
    actual_result += key["Words"][0]["WordText"]
print (actual_result)

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

expected_result = sheet.cell(row=2, column=1).value
sheet['B2'] = actual_result
print (sheet.cell(row=2, column=2).value)
if  expected_result == actual_result:
   sheet.cell(row=2, column=3).value = 'Pass'
else:
   sheet.cell(row=2, column=3).value = 'Fail'
wb.save('example.xlsx')

folder_name = "backup"
os.makedirs(folder_name, exist_ok=True)
shutil.move( file_name, folder_name + "/" + file_name)

