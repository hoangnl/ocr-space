import openpyxl


wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print (sheet['A2'].value)

print (sheet.cell(row=2, column=1).value)
sheet['B2'] = 'B'
print (sheet.cell(row=2, column=2).value)
if sheet.cell(row=2, column=1).value == sheet.cell(row=2, column=2).value:
   sheet['C2'] = 'Pass'
else:
   sheet['C2'] = 'Fail'
wb.save('example.xlsx');